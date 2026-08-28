# ------------------------------------------------------------------
# Copyright (c) Qualcomm Technologies, Inc. and/or its subsidiaries.
# SPDX-License-Identifier: BSD-3-Clause-Clear
# ------------------------------------------------------------------
"""
Parse all parse_<debug>_<model>_tp<N>.log files in a given directory and produce
an Excel summary.

Usage:
    python parse_logs_to_excel.py --log-dir <path/to/logs> [--date "Apr 21 2026"]

If --log-dir is omitted the script's own directory is used.
If --date is provided only log files whose second line contains that date are parsed.
Accepted date formats: "Apr 21 2026", "Apr 21", "2026-04-21", "21 Apr 2026"

Every TP size and debug flag is included by default; the TP a model actually ran
with is reported in a per-model "TP" column. Use --tp-size / --qaic-debug to
narrow the selection.

The output Excel file is written to <log-dir>/fallback_ops_summary.xlsx.
"""

import argparse
import ast
import glob
import json
import os
from datetime import datetime

import pandas as pd
import regex as re

# Date formats accepted from the user via --date
_DATE_INPUT_FORMATS = [
    "%b %d %Y",  # Apr 21 2026
    "%b %d",  # Apr 21  (no year)
    "%Y-%m-%d",  # 2026-04-21
    "%d %b %Y",  # 21 Apr 2026
    "%d %b",  # 21 Apr  (no year)
]

# Format used inside the log file second line: "Tue Apr 21 21:36:54 2026"
_LOG_DATE_FORMAT = "%a %b %d %H:%M:%S %Y"

# parse_<debug>_<model>_tp<N>.log — the tp suffix is the *effective* TP the model
# ran with, which may differ from the sweep's --tp-size when
# model_configs_<type>.py pins tp_size for that model. The debug group tolerates
# an empty slot, since sourcing unset_qaic_debug.sh can leave it blank.
_PARSE_NAME_RE = re.compile(r"^parse_(?P<debug>\d*)_(?P<model>.+)_tp(?P<tp>\d+)\.log$")


def parse_user_date(date_str):
    """Parse the user-supplied --date string into a (datetime.date, has_year) tuple.

    has_year is True when the user included a year, False otherwise (only
    month+day will be compared against the log file date).
    """
    date_str = date_str.strip()
    for fmt in _DATE_INPUT_FORMATS:
        try:
            dt = datetime.strptime(date_str, fmt)
            has_year = "%Y" in fmt
            return dt.date(), has_year
        except ValueError:
            continue
    raise ValueError(
        f"Unrecognised date format: '{date_str}'.\n"
        "Accepted formats: 'Apr 21 2026', 'Apr 21', '2026-04-21', '21 Apr 2026'"
    )


def get_log_date(log_file):
    """Read the second line of a log file and return its date as datetime.date, or None.

    Expected second-line format: Tue Apr 21 21:36:54 2026
    """
    try:
        with open(log_file) as f:
            for i, line in enumerate(f):
                if i == 1:  # zero-indexed: line index 1 == second line
                    line = line.strip()
                    try:
                        dt = datetime.strptime(line, _LOG_DATE_FORMAT)
                        return dt.date()
                    except ValueError:
                        # Fallback: extract date parts via regex in case of
                        # extra whitespace or a slightly different prefix
                        m = re.search(
                            r"(\w{3})\s+(\w{3})\s+(\d{1,2})\s+\d{2}:\d{2}:\d{2}\s+(\d{4})",
                            line,
                        )
                        if m:
                            rebuilt = (
                                f"{m.group(1)} {m.group(2)} "
                                f"{int(m.group(3)):2d} 00:00:00 {m.group(4)}"
                            )
                            try:
                                return datetime.strptime(
                                    rebuilt, _LOG_DATE_FORMAT
                                ).date()
                            except ValueError:
                                pass
                    return None
    except OSError:
        pass
    return None


def parse_log_file(log_file):
    """Extract the ops dict from a parse_<debug>_<model>_tp<N>.log file."""
    with open(log_file) as f:
        content = f.read()

    # Match the dict portion inside defaultdict(..., {...})
    match = re.search(r"defaultdict\([^,]+,\s*(\{.*\})\)", content)
    if match:
        dict_str = match.group(1)
        try:
            return ast.literal_eval(dict_str)
        except Exception as e:
            raise ValueError(f"Could not parse dict in {log_file}") from e
            return {}
    return {}


def check_model_ran(log_dir, parse_log_basename):
    """Return 'Yes' if the corresponding log_<debug>_* file shows
    output was generated, else 'No'.

    Looks for 'Processed prompts: 100%' which only appears when inference completed
    and tokens were generated.
    """
    # parse_<debug>_<model>_tp<N>.log -> log_<debug>_<model>_tp<N>.log.
    # count=1 so only the leading prefix is rewritten, never a model name.
    orig_basename = parse_log_basename.replace("parse_", "log_", 1)
    orig_log = os.path.join(log_dir, orig_basename)
    if not os.path.exists(orig_log):
        return "N/A"
    try:
        with open(orig_log) as f:
            for line in f:
                if "Processed prompts: 100%" in line:
                    return "Yes"
    except OSError:
        return "N/A"
    return "No"


def build_model_type_map(models_dir, ref=None):
    """
    Read the <type>_models_<ref>.json files in models_dir and return a dict
    mapping sanitised model name (slashes→underscores) to its type string.
    E.g. {"Qwen_Qwen3-8B": "llm", "OpenGVLab_InternVL3-9B": "vlm"}

    scrape_models.py stamps each file with the vLLM ref it was scraped from.
    When ref is given only that ref's files are read; otherwise every ref found
    is merged, which is safe because a model's type does not vary by ref.
    """
    model_type_map = {}
    if not models_dir:
        return model_type_map
    # Mirror scrape_models.py's ref_suffix(): '/' is not filename-safe
    suffix = ref.replace("/", "_") if ref else "*"
    pattern = os.path.join(models_dir, f"*_models_{suffix}.json")
    matches = sorted(glob.glob(pattern))
    if not matches:
        print(f"  Warning: no model list matched {pattern} — using a single sheet")
    for json_file in matches:
        basename = os.path.basename(json_file)
        # "llm_models_v0.23.0.json" -> "llm"; also tolerates "llm_models.json"
        model_type = basename.split("_models")[0]
        try:
            with open(json_file) as f:
                entries = json.load(f)
            for e in entries:
                sanitised = e["model"].replace("/", "_").replace(" ", "_")
                model_type_map[sanitised] = model_type
        except Exception as exc:
            print(f"  Warning: could not read {json_file}: {exc}")
    return model_type_map


def build_sheet(writer, sheet_name, records):
    """Write one sheet to the Excel writer for the given records.

    records is a list of dicts with keys: model, tp, ops, ran. A model may appear
    more than once when it was run at several TP sizes, so rows are keyed by
    (model, tp) rather than by model alone.
    """
    all_ops = sorted({op for r in records for op in r["ops"]})

    if not all_ops:
        print(f"  No operations found for sheet '{sheet_name}'. Writing empty sheet.")

    # Excel-facing labels for the per-op sub-columns. The underlying dict keys
    # ("tc", "tt") come from fallback_parser.py's printed output and are kept
    # as-is so existing parse_*.log files still parse correctly.
    SUB_COLUMNS = [("Fallback Count", "tc"), ("Total Time (us)", "tt")]

    columns = pd.MultiIndex.from_tuples(
        [("Ran?", ""), ("Fallback?", ""), ("TP", "")]
        + [(op, label) for op in all_ops for label, _ in SUB_COLUMNS],
        names=["Operation", "Metric"],
    )

    records = sorted(records, key=lambda r: (r["model"], r["tp"]))
    index = [r["model"] for r in records]
    rows = []
    for r in records:
        fallback = "Yes" if r["ops"] else "No"
        row = [r["ran"], fallback, r["tp"]]
        for op in all_ops:
            for _, key in SUB_COLUMNS:
                row.append(r["ops"].get(op, {}).get(key, ""))
        rows.append(row)

    df = pd.DataFrame(rows, index=index, columns=columns)
    df.index.name = "Model"
    df.to_excel(writer, sheet_name=sheet_name)

    # Auto-fit column widths
    from openpyxl.cell.cell import MergedCell
    from openpyxl.utils import get_column_letter

    ws = writer.sheets[sheet_name]
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[col_letter]:
            if not isinstance(cell, MergedCell) and cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 45)


def build_excel(
    log_dir,
    output_file,
    filter_date=None,
    filter_has_year=True,
    models_dir=None,
    ref=None,
    tp_filter=None,
    debug_filter=None,
):
    # Match any debug flag and any TP. The effective TP varies per model when
    # model_configs_<type>.py pins tp_size, so it cannot be hardcoded here --
    # doing so silently drops those runs from the spreadsheet.
    log_files = sorted(glob.glob(os.path.join(log_dir, "parse_*_tp*.log")))
    if not log_files:
        print(f"No parse_*_tp*.log files found in {log_dir}")
        return

    # Date filtering
    if filter_date is not None:
        filtered = []
        for lf in log_files:
            log_date = get_log_date(lf)
            if log_date is None:
                print(f"  Skipping (no date in second line): {os.path.basename(lf)}")
                continue
            if filter_has_year:
                match = log_date == filter_date
            else:
                match = (
                    log_date.month == filter_date.month
                    and log_date.day == filter_date.day
                )
            if match:
                filtered.append(lf)
            else:
                print(
                    f"  Skipping (date {log_date} != {filter_date}): "
                    f"{os.path.basename(lf)}"
                )
        log_files = filtered
        if not log_files:
            print(f"No log files matched date filter '{filter_date}'.")
            return

    # Build model→type map from JSON files if provided; empty map = single sheet mode
    model_type_map = build_model_type_map(models_dir, ref)

    # Parse all log files, group by type (or single bucket if no map provided)
    sheets = {}  # sheet_name -> [record, ...]
    skipped = 0

    for log_file in log_files:
        basename = os.path.basename(log_file)
        m = _PARSE_NAME_RE.match(basename)
        if not m:
            print(f"  Skipping (unrecognised name): {basename}")
            skipped += 1
            continue

        model_name = m.group("model")
        tp = int(m.group("tp"))
        debug = m.group("debug")

        if tp_filter is not None and tp != tp_filter:
            continue
        if debug_filter is not None and debug != str(debug_filter):
            continue

        ops_dict = parse_log_file(log_file)
        ran = check_model_ran(log_dir, basename)
        print(f"  {model_name} (tp{tp}): {len(ops_dict)} ops  |  ran={ran}")

        # If a type map exists use it; otherwise fall back to single sheet
        sheet = (
            model_type_map.get(model_name, "Fallback Ops")
            if model_type_map
            else "Fallback Ops"
        )
        if model_type_map and sheet != "Fallback Ops":
            sheet = sheet.upper()
        sheets.setdefault(sheet, []).append(
            {"model": model_name, "tp": tp, "ops": ops_dict, "ran": ran}
        )

    if skipped:
        print(f"  ({skipped} file(s) skipped for unrecognised names)")

    if not sheets:
        print("No log files left after filtering — nothing to write.")
        return

    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        total_models = 0
        for sheet_name in sorted(sheets.keys()):
            build_sheet(writer, sheet_name, sheets[sheet_name])
            total_models += len(sheets[sheet_name])
            print(f"  Sheet '{sheet_name}': {len(sheets[sheet_name])} models")

    print(f"\nExcel written to: {output_file}")
    print(f"Total models: {total_models}  |  Sheets: {list(sheets.keys())}")


def main():
    parser = argparse.ArgumentParser(
        description=(
            "Parse parse_<debug>_<model>_tp<N>.log files and produce an Excel summary."
        )
    )
    parser.add_argument(
        "--log-dir",
        default=os.path.dirname(os.path.abspath(__file__)),
        help=(
            "Directory containing parse_<debug>_<model>_tp<N>.log files "
            "(default: script directory)"
        ),
    )
    parser.add_argument(
        "--output",
        default=None,
        help="Output Excel file path (default: <log-dir>/fallback_ops_summary.xlsx)",
    )
    parser.add_argument(
        "--date",
        default=None,
        metavar="DATE",
        help=(
            "Only parse log files whose second line contains this date. "
            "Accepted formats: 'Apr 21 2026', 'Apr 21', '2026-04-21', '21 Apr 2026'"
        ),
    )
    parser.add_argument(
        "--models-dir",
        default=None,
        help="Directory with <type>_models_<ref>.json files — enables per-type sheets",
    )
    parser.add_argument(
        "--ref",
        default=None,
        metavar="GIT_REF",
        help=(
            "Only read the model lists scraped from this vLLM ref, i.e. "
            "<type>_models_<ref>.json (default: merge every ref found)"
        ),
    )
    parser.add_argument(
        "--tp-size",
        type=int,
        default=None,
        metavar="N",
        help="Only include logs whose effective TP is N (default: all TP sizes)",
    )
    parser.add_argument(
        "--qaic-debug",
        type=int,
        default=None,
        metavar="0|1",
        help="Only include logs with this QAIC_DEBUG flag (default: all)",
    )
    args = parser.parse_args()

    log_dir = os.path.abspath(args.log_dir)
    output_file = args.output or os.path.join(log_dir, "fallback_ops_summary.xlsx")

    filter_date = None
    filter_has_year = True
    if args.date:
        try:
            filter_date, filter_has_year = parse_user_date(args.date)
        except ValueError as exc:
            print(f"Error: {exc}")
            raise SystemExit(1) from exc

    print(f"Log directory : {log_dir}")
    print(f"Output file   : {output_file}")
    if args.models_dir:
        print(f"Models dir    : {args.models_dir}")
    if args.ref:
        print(f"Model list ref: {args.ref}")
    if args.tp_size is not None:
        print(f"TP filter     : {args.tp_size}")
    if args.qaic_debug is not None:
        print(f"Debug filter  : {args.qaic_debug}")
    if filter_date:
        year_info = str(filter_date.year) if filter_has_year else "(any year)"
        print(f"Date filter   : {filter_date.strftime('%b %d')} {year_info}")
    print()

    build_excel(
        log_dir,
        output_file,
        filter_date=filter_date,
        filter_has_year=filter_has_year,
        models_dir=args.models_dir,
        ref=args.ref,
        tp_filter=args.tp_size,
        debug_filter=args.qaic_debug,
    )


if __name__ == "__main__":
    main()
